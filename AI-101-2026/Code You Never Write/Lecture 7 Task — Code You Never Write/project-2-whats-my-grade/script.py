#!/usr/bin/env python3
"""
grade_calculator.py
--------------------
Reusable, commandline-runnable grade calculator for the Course Grade Tracker
workbook format (a "Weight table" in columns A:D and an "Assignment table"
in columns F:J on a sheet named 'Tracker').

Usage:
    python3 grade_calculator.py <path_to_xlsx> [--sheet Tracker] [--target 90]

Logic overview (see README section in the accompanying report for full detail):
  1. Parse the weight table (Category Name, Weight) from columns A:B.
  2. Parse the assignment table (Category, Assignment Name, Score, Max Points)
     from columns F:I. Any row that can't be parsed (missing category, missing
     max points, non-numeric score/max) is skipped and flagged.
  3. Per the task's edge-case rule, any assignment with a missing/blank Score
     is treated as a score of 0 (not excluded).
  4. Category average:
       - For 'Quizzes' specifically, the workbook's own formula encodes a
         "drop the lowest quiz" policy. We replicate that: drop the row with
         the minimum score and use (sum(scores)-min_score) / (sum(max)-min(max))
       - All other categories: simple average of each assignment's own
         percentage (score/max), which is how the workbook's AVERAGEIF-based
         formula computes it.
  5. Overall grade = sum(weight_c * category_avg_c) across ALL categories
     (weights always sum to 1, missing work counts as 0 -- no reweighting).
  6. Letter grade from the fixed grading table (A/B/C/D/F).
  7. Minimum ordered steps to reach a target grade (default: 90, i.e. an A):
       - If already at/above target: 0 steps.
       - Else, any category with a missing (not-yet-taken) assignment is a
         "pending" step: solve for the score needed on that assignment (in
         the order the pending assignments appear) so that, combined with
         everything else held fixed, the target is reached -- capped at the
         assignment's max points.
       - If pending assignments alone can't close the gap, additional steps
         are added by improving the *already completed* category that gives
         the most overall percentage-point gain per additional raw point
         (i.e., weight / category_max_points), in descending order of that
         efficiency, until the target is met.
"""

import argparse
import copy
import json
import sys

from openpyxl import load_workbook

GRADE_BANDS = [
    (90, 100.0000001, "A"),
    (80, 90, "B"),
    (70, 80, "C"),
    (60, 70, "D"),
    (0, 60, "F"),
]


def letter_grade(pct):
    for lo, hi, letter in GRADE_BANDS:
        if lo <= pct < hi:
            return letter
    return "F"


def parse_workbook(path, sheet_name):
    wb = load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")
    ws = wb[sheet_name]

    weights = {}
    flagged_rows = []

    # Weight table: columns A (Category Name) & B (Weight), starting row 2.
    row_idx = 2
    while True:
        name = ws.cell(row=row_idx, column=1).value
        weight = ws.cell(row=row_idx, column=2).value
        if name is None:
            break
        if str(name).strip().lower().startswith("total"):
            row_idx += 1
            continue
        try:
            weights[str(name).strip()] = float(weight)
        except (TypeError, ValueError):
            flagged_rows.append({"table": "weights", "row": row_idx, "reason": f"Non-numeric weight: {weight!r}"})
        row_idx += 1

    # Assignment table: columns F (Category), G (Assignment Name), H (Score), I (Max Points)
    assignments = []
    row_idx = 2
    while True:
        category = ws.cell(row=row_idx, column=6).value
        name = ws.cell(row=row_idx, column=7).value
        score = ws.cell(row=row_idx, column=8).value
        max_points = ws.cell(row=row_idx, column=9).value
        if category is None and name is None and max_points is None:
            break

        if category is None or max_points is None:
            flagged_rows.append({
                "table": "assignments", "row": row_idx,
                "reason": f"Missing category or max points (category={category!r}, max_points={max_points!r})"
            })
            row_idx += 1
            continue

        try:
            max_points = float(max_points)
        except (TypeError, ValueError):
            flagged_rows.append({
                "table": "assignments", "row": row_idx,
                "reason": f"Non-numeric max points: {max_points!r}"
            })
            row_idx += 1
            continue

        was_missing = score is None or (isinstance(score, str) and score.strip() == "")
        if was_missing:
            score_val = 0.0
        else:
            try:
                score_val = float(score)
            except (TypeError, ValueError):
                flagged_rows.append({
                    "table": "assignments", "row": row_idx,
                    "reason": f"Non-numeric score: {score!r}"
                })
                row_idx += 1
                continue

        assignments.append({
            "category": str(category).strip(),
            "name": name,
            "score": score_val,
            "max_points": max_points,
            "was_missing": was_missing,
            "row": row_idx,
        })
        row_idx += 1

    return weights, assignments, flagged_rows


def compute_category_average(category, items):
    """Returns (category_avg_pct, detail_dict)."""
    if category.strip().lower() == "quizzes" and len(items) > 1:
        scores = [it["score"] for it in items]
        maxes = [it["max_points"] for it in items]
        min_score = min(scores)
        min_max = min(maxes)
        denom = sum(maxes) - min_max
        numer = sum(scores) - min_score
        avg = (numer / denom * 100) if denom else 0.0
        detail = {
            "method": "drop_lowest_quiz",
            "sum_scores": sum(scores), "min_score_dropped": min_score,
            "sum_max": sum(maxes), "min_max_dropped": min_max,
            "formula": f"({sum(scores)} - {min_score}) / ({sum(maxes)} - {min_max}) * 100",
        }
        return avg, detail
    else:
        pct_list = [(it["score"] / it["max_points"] * 100) if it["max_points"] else 0.0 for it in items]
        avg = sum(pct_list) / len(pct_list) if pct_list else 0.0
        detail = {
            "method": "simple_average_of_percentages",
            "per_assignment_pct": [round(p, 4) for p in pct_list],
        }
        return avg, detail


def compute_grade(weights, assignments):
    by_category = {}
    for a in assignments:
        by_category.setdefault(a["category"], []).append(a)

    category_results = {}
    for category, weight in weights.items():
        items = by_category.get(category, [])
        if not items:
            category_results[category] = {
                "weight": weight, "average_pct": 0.0, "items": [], "detail": {"method": "no_data"}
            }
            continue
        avg, detail = compute_category_average(category, items)
        category_results[category] = {"weight": weight, "average_pct": avg, "items": items, "detail": detail}

    overall = sum(c["weight"] * c["average_pct"] for c in category_results.values())
    return overall, category_results


def efficiency_plan(weights, category_results, target):
    """Determine ordered list of minimum steps to reach `target` overall pct."""
    overall = sum(c["weight"] * c["average_pct"] for c in category_results.values())
    steps = []
    if overall >= target:
        return steps, overall

    # 1) Pending (missing) assignments first, in the order they appear.
    pending = []
    for category, res in category_results.items():
        for item in res["items"]:
            if item["was_missing"]:
                pending.append((category, item))

    running_overall = overall
    for category, item in pending:
        if running_overall >= target:
            break
        res = category_results[category]
        weight = res["weight"]
        others_contribution = running_overall - weight * res["average_pct"]
        gap = target - others_contribution
        # required category average pct to close the gap on its own
        required_cat_avg = gap / weight if weight else float("inf")
        # figure out what raw score on this single missing item achieves that,
        # given the category's averaging method holds other items constant.
        items = res["items"]
        if category.strip().lower() == "quizzes" and len(items) > 1:
            other_scores = [it["score"] for it in items if it is not item]
            other_maxes = [it["max_points"] for it in items if it is not item]
            # replicate drop-lowest logic: try assuming this item's score is not the min after change
            # solve simple linear system: new_avg = (sum(other_scores)+x - min(scores_all)) / (sum(other_maxes)+max - min(maxes_all)) *100
            # For a 2-quiz case (this dataset), this reduces to needing x alone since the other becomes the dropped one.
            best_x = None
            max_pts = item["max_points"]
            for x in [v / 4 for v in range(0, int(max_pts * 4) + 1)]:
                trial_scores = other_scores + [x]
                trial_maxes = other_maxes + [max_pts]
                min_s, min_m = min(trial_scores), min(trial_maxes)
                denom = sum(trial_maxes) - min_m
                cat_avg = (sum(trial_scores) - min_s) / denom * 100 if denom else 0.0
                if cat_avg + 1e-9 >= required_cat_avg:
                    best_x = x
                    break
            required_score = best_x
        else:
            required_score = required_cat_avg / 100 * item["max_points"]

        capped = False
        if required_score is None or required_score > item["max_points"]:
            required_score = item["max_points"]
            capped = True
        required_score = max(0.0, required_score)

        # apply step
        item["score"] = required_score
        item["was_missing"] = False
        new_avg, detail = compute_category_average(category, items)
        res["average_pct"] = new_avg
        res["detail"] = detail
        new_overall = sum(c["weight"] * c["average_pct"] for c in category_results.values())

        steps.append({
            "type": "complete_pending_assignment",
            "category": category,
            "assignment": item["name"],
            "required_score": round(required_score, 2),
            "max_points": item["max_points"],
            "resulting_overall_pct": round(new_overall, 4),
            "capped_at_max": capped,
        })
        running_overall = new_overall

    if running_overall >= target:
        return steps, running_overall

    # 2) Improve already-completed categories, most-efficient (weight / category_max) first.
    ordering = []
    for category, res in category_results.items():
        max_pts = sum(it["max_points"] for it in res["items"]) if res["items"] else 0
        if max_pts <= 0:
            continue
        efficiency = res["weight"] / max_pts
        ordering.append((efficiency, category))
    ordering.sort(reverse=True)

    for efficiency, category in ordering:
        if running_overall >= target:
            break
        res = category_results[category]
        weight = res["weight"]
        others_contribution = running_overall - weight * res["average_pct"]
        gap = target - others_contribution
        required_cat_avg = gap / weight if weight else float("inf")
        current_avg = res["average_pct"]
        if required_cat_avg <= current_avg:
            continue
        total_max = sum(it["max_points"] for it in res["items"])
        total_current_score = sum(it["score"] for it in res["items"])
        required_total_score = required_cat_avg / 100 * total_max
        extra_points_needed = required_total_score - total_current_score
        capped = False
        if required_total_score > total_max:
            required_total_score = total_max
            extra_points_needed = total_max - total_current_score
            capped = True

        # distribute the extra points onto the single largest-headroom assignment for simplicity
        target_item = max(res["items"], key=lambda it: it["max_points"] - it["score"])
        target_item["score"] = min(target_item["max_points"], target_item["score"] + extra_points_needed)
        new_avg, detail = compute_category_average(category, res["items"])
        res["average_pct"] = new_avg
        res["detail"] = detail
        new_overall = sum(c["weight"] * c["average_pct"] for c in category_results.values())

        steps.append({
            "type": "improve_existing_category",
            "category": category,
            "assignment": target_item["name"],
            "extra_points_needed": round(extra_points_needed, 2),
            "new_score": round(target_item["score"], 2),
            "max_points": target_item["max_points"],
            "resulting_overall_pct": round(new_overall, 4),
            "capped_at_max": capped,
        })
        running_overall = new_overall

    return steps, running_overall


def main():
    parser = argparse.ArgumentParser(description="Compute course grade from a Grade Tracker workbook.")
    parser.add_argument("xlsx_path")
    parser.add_argument("--sheet", default="Tracker")
    parser.add_argument("--target", type=float, default=90.0, help="Target overall percentage (default 90 = A).")
    args = parser.parse_args()

    weights, assignments, flagged_rows = parse_workbook(args.xlsx_path, args.sheet)
    overall, category_results = compute_grade(weights, assignments)
    grade = letter_grade(overall)

    hw2 = next((a for a in assignments if a["name"] == "HW 2"), None)
    hw2_check = None
    if hw2:
        hw2_check = {
            "score": hw2["score"], "max_points": hw2["max_points"],
            "percentage": round(hw2["score"] / hw2["max_points"] * 100, 4),
        }

    # Snapshot CURRENT state before the planner mutates scores to explore the path to target.
    current_category_averages = {c: round(r["average_pct"], 4) for c, r in category_results.items()}
    current_category_detail = {c: copy.deepcopy(r["detail"]) for c, r in category_results.items()}

    plan_category_results = copy.deepcopy(category_results)
    steps, final_overall = efficiency_plan(weights, plan_category_results, args.target)

    result = {
        "weights": weights,
        "category_averages": current_category_averages,
        "category_detail": current_category_detail,
        "overall_pct": round(overall, 4),
        "letter_grade": grade,
        "target_pct": args.target,
        "hw2_manual_check": hw2_check,
        "flagged_rows": flagged_rows,
        "min_steps_to_target": steps,
        "num_steps_to_target": len(steps),
        "projected_overall_after_steps": round(final_overall, 4),
    }
    print(json.dumps(result, indent=2, default=str))


if __name__ == "__main__":
    main()
